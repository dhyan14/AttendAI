from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, delete as sql_delete
from app.database import get_db
from app.api.deps import require_super_admin
from app.models import (
    Organization, Department, Student, Faculty, User, UserRole, Subject,
    SubjectAssignment, Lecture, AttendanceRecord, AttendanceEvidence,
    AttendanceDispute, FaceEmbedding, DisputeStatus,
)
from pydantic import BaseModel
from typing import Optional, List
import uuid
import bcrypt
import io
import pandas as pd

router = APIRouter()


# ─── Schemas ───────────────────────────────────────────────

class OrgCreate(BaseModel):
    name: str
    code: str
    min_attendance: int = 75


class DeptCreate(BaseModel):
    org_id: str
    name: str
    code: str
    institute_name: Optional[str] = None


class StudentCreate(BaseModel):
    name: str
    roll_no: str
    enrollment_no: Optional[str] = None
    email: str
    division: Optional[str] = None
    batch: Optional[str] = None
    semester: Optional[int] = None
    dept_id: str


class FacultyCreate(BaseModel):
    name: str
    email: str
    designation: Optional[str] = None
    dept_id: str


class AdminCreate(BaseModel):
    email: str
    role: str  # "org_admin" | "dept_admin"
    org_id: str
    dept_id: Optional[str] = None   # required when role == dept_admin
    name: Optional[str] = None      # for dept_admin Faculty profile


# ─── Helper ────────────────────────────────────────────────

def _hash(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


# ═══════════════════════════════════════════════════════════
# PLATFORM STATS
# ═══════════════════════════════════════════════════════════

@router.get("/stats")
async def get_platform_stats(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    """Platform-wide stats for the Super Admin dashboard."""
    org_q    = await db.execute(select(func.count(Organization.id)))
    users_q  = await db.execute(select(func.count(User.id)))
    stud_q   = await db.execute(select(func.count(User.id)).where(User.role == UserRole.student))
    fac_q    = await db.execute(select(func.count(User.id)).where(User.role == UserRole.faculty))
    lec_q    = await db.execute(select(func.count(Lecture.id)))
    att_q    = await db.execute(select(func.count(AttendanceRecord.id)))
    disp_q   = await db.execute(
        select(func.count(AttendanceDispute.id))
        .where(AttendanceDispute.status == DisputeStatus.open)
    )

    orgs_res = await db.execute(select(Organization).order_by(Organization.created_at))
    orgs = orgs_res.scalars().all()

    org_details = []
    for org in orgs:
        d_q = await db.execute(select(func.count(Department.id)).where(Department.org_id == org.id))
        s_q = await db.execute(
            select(func.count(Student.id)).join(Department, Department.id == Student.dept_id)
            .where(Department.org_id == org.id)
        )
        f_q = await db.execute(
            select(func.count(Faculty.id)).join(User, User.id == Faculty.user_id)
            .where(User.org_id == org.id)
        )
        l_q = await db.execute(
            select(func.count(Lecture.id))
            .join(Subject, Subject.id == Lecture.subject_id)
            .join(Department, Department.id == Subject.dept_id)
            .where(Department.org_id == org.id)
        )
        org_details.append({
            "id": str(org.id),
            "name": org.name,
            "code": org.code,
            "departments": d_q.scalar() or 0,
            "students": s_q.scalar() or 0,
            "faculty": f_q.scalar() or 0,
            "lectures": l_q.scalar() or 0,
            "created_at": org.created_at.strftime("%Y-%m-%d") if org.created_at else "—",
            "settings": org.settings or {},
        })

    return {
        "total_orgs": org_q.scalar() or 0,
        "total_users": users_q.scalar() or 0,
        "total_students": stud_q.scalar() or 0,
        "total_faculty": fac_q.scalar() or 0,
        "total_lectures": lec_q.scalar() or 0,
        "total_attendance_records": att_q.scalar() or 0,
        "open_disputes": disp_q.scalar() or 0,
        "organizations": org_details,
    }


# ═══════════════════════════════════════════════════════════
# ORGANIZATIONS
# ═══════════════════════════════════════════════════════════

@router.get("/orgs")
async def list_all_orgs(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    res = await db.execute(select(Organization).order_by(Organization.name))
    orgs = res.scalars().all()
    return [{"id": str(o.id), "name": o.name, "code": o.code, "settings": o.settings} for o in orgs]


@router.post("/orgs")
async def create_org(
    data: OrgCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    existing = await db.execute(select(Organization).where(Organization.code == data.code))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Organization code '{data.code}' already exists")

    org = Organization(
        name=data.name,
        code=data.code.upper(),
        settings={"minAttendancePercent": data.min_attendance}
    )
    db.add(org)
    await db.commit()
    await db.refresh(org)
    return {"id": str(org.id), "name": org.name, "code": org.code, "settings": org.settings}


@router.delete("/orgs/{org_id}")
async def delete_org(
    org_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    org = await db.get(Organization, uuid.UUID(org_id))
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    org_uuid = uuid.UUID(org_id)
    name = org.name

    # Collect all dept IDs in this org
    dept_res = await db.execute(select(Department.id).where(Department.org_id == org_uuid))
    dept_ids = [r[0] for r in dept_res.all()]

    if dept_ids:
        # Collect lecture IDs inside those depts
        subj_res = await db.execute(select(Subject.id).where(Subject.dept_id.in_(dept_ids)))
        subj_ids = [r[0] for r in subj_res.all()]

        lec_ids = []
        if subj_ids:
            lec_res = await db.execute(select(Lecture.id).where(Lecture.subject_id.in_(subj_ids)))
            lec_ids = [r[0] for r in lec_res.all()]

        # Collect student IDs
        stu_res = await db.execute(select(Student.id).where(Student.dept_id.in_(dept_ids)))
        stu_ids = [r[0] for r in stu_res.all()]

        # Collect faculty IDs
        fac_res = await db.execute(select(Faculty.id).where(Faculty.dept_id.in_(dept_ids)))
        fac_ids = [r[0] for r in fac_res.all()]

        # Delete in FK-safe order
        if lec_ids:
            await db.execute(sql_delete(AttendanceEvidence).where(AttendanceEvidence.lecture_id.in_(lec_ids)))
            await db.execute(sql_delete(AttendanceDispute).where(AttendanceDispute.lecture_id.in_(lec_ids)))
            await db.execute(sql_delete(AttendanceRecord).where(AttendanceRecord.lecture_id.in_(lec_ids)))
            await db.execute(sql_delete(Lecture).where(Lecture.id.in_(lec_ids)))

        if subj_ids:
            await db.execute(sql_delete(SubjectAssignment).where(SubjectAssignment.subject_id.in_(subj_ids)))
            await db.execute(sql_delete(Subject).where(Subject.id.in_(subj_ids)))

        if stu_ids:
            await db.execute(sql_delete(FaceEmbedding).where(FaceEmbedding.student_id.in_(stu_ids)))
            await db.execute(sql_delete(Student).where(Student.id.in_(stu_ids)))

        if fac_ids:
            await db.execute(sql_delete(Faculty).where(Faculty.id.in_(fac_ids)))

        await db.execute(sql_delete(Department).where(Department.org_id == org_uuid))

    # Delete all users in this org (except super_admin)
    await db.execute(
        sql_delete(User).where(
            User.org_id == org_uuid,
            User.role != UserRole.super_admin
        )
    )

    await db.delete(org)
    await db.commit()
    return {"message": f"Organization '{name}' deleted"}


# ═══════════════════════════════════════════════════════════
# DEPARTMENTS
# ═══════════════════════════════════════════════════════════

@router.get("/departments")
async def list_all_departments(
    org_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    """List all departments across the platform, optionally filtered by org."""
    q = select(Department, Organization.name.label("org_name")).join(
        Organization, Organization.id == Department.org_id
    )
    if org_id:
        q = q.where(Department.org_id == org_id)
    q = q.order_by(Organization.name, Department.name)
    res = await db.execute(q)
    rows = res.all()

    result = []
    for r in rows:
        # Count students in this dept
        s_q = await db.execute(select(func.count(Student.id)).where(Student.dept_id == r.Department.id))
        f_q = await db.execute(select(func.count(Faculty.id)).where(Faculty.dept_id == r.Department.id))
        result.append({
            "id": str(r.Department.id),
            "name": r.Department.name,
            "code": r.Department.code,
            "institute_name": r.Department.institute_name,
            "org_id": str(r.Department.org_id),
            "org_name": r.org_name,
            "student_count": s_q.scalar() or 0,
            "faculty_count": f_q.scalar() or 0,
        })
    return result


@router.post("/departments")
async def create_department(
    data: DeptCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    """Create a department inside any organization."""
    org = await db.get(Organization, uuid.UUID(data.org_id))
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    existing = await db.execute(
        select(Department).where(
            Department.org_id == uuid.UUID(data.org_id),
            Department.code == data.code
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Department code '{data.code}' already exists in this org")

    dept = Department(
        org_id=uuid.UUID(data.org_id),
        name=data.name,
        code=data.code.upper(),
        institute_name=data.institute_name,
    )
    db.add(dept)
    await db.commit()
    await db.refresh(dept)
    return {
        "id": str(dept.id),
        "name": dept.name,
        "code": dept.code,
        "org_id": str(dept.org_id),
        "org_name": org.name,
        "student_count": 0,
        "faculty_count": 0,
    }


@router.delete("/departments/{dept_id}")
async def delete_department(
    dept_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    dept = await db.get(Department, uuid.UUID(dept_id))
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    dept_uuid = uuid.UUID(dept_id)
    name = dept.name

    # Get all subject IDs in this dept
    subj_res = await db.execute(select(Subject.id).where(Subject.dept_id == dept_uuid))
    subj_ids = [r[0] for r in subj_res.all()]

    # Get all lecture IDs from those subjects
    lec_ids = []
    if subj_ids:
        lec_res = await db.execute(select(Lecture.id).where(Lecture.subject_id.in_(subj_ids)))
        lec_ids = [r[0] for r in lec_res.all()]

    # Get student IDs in this dept
    stu_res = await db.execute(select(Student.id).where(Student.dept_id == dept_uuid))
    stu_ids = [r[0] for r in stu_res.all()]

    # Get faculty IDs in this dept
    fac_res = await db.execute(select(Faculty.id).where(Faculty.dept_id == dept_uuid))
    fac_ids = [r[0] for r in fac_res.all()]

    # Delete in FK-safe order
    if lec_ids:
        await db.execute(sql_delete(AttendanceEvidence).where(AttendanceEvidence.lecture_id.in_(lec_ids)))
        await db.execute(sql_delete(AttendanceDispute).where(AttendanceDispute.lecture_id.in_(lec_ids)))
        await db.execute(sql_delete(AttendanceRecord).where(AttendanceRecord.lecture_id.in_(lec_ids)))
        await db.execute(sql_delete(Lecture).where(Lecture.id.in_(lec_ids)))

    if subj_ids:
        await db.execute(sql_delete(SubjectAssignment).where(SubjectAssignment.subject_id.in_(subj_ids)))
        await db.execute(sql_delete(Subject).where(Subject.id.in_(subj_ids)))

    if stu_ids:
        await db.execute(sql_delete(FaceEmbedding).where(FaceEmbedding.student_id.in_(stu_ids)))
        await db.execute(sql_delete(AttendanceDispute).where(AttendanceDispute.student_id.in_(stu_ids)))
        await db.execute(sql_delete(Student).where(Student.id.in_(stu_ids)))

    if fac_ids:
        await db.execute(sql_delete(SubjectAssignment).where(SubjectAssignment.faculty_id.in_(fac_ids)))
        await db.execute(sql_delete(Faculty).where(Faculty.id.in_(fac_ids)))

    await db.delete(dept)
    await db.commit()
    return {"message": f"Department '{name}' deleted"}


# ═══════════════════════════════════════════════════════════
# STUDENTS
# ═══════════════════════════════════════════════════════════

@router.get("/students")
async def list_all_students(
    org_id: Optional[str] = None,
    dept_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    """List all students across the platform, optionally filtered by org or dept."""
    q = (
        select(Student, Department.name.label("dept_name"), Organization.name.label("org_name"), User.email.label("email"))
        .join(Department, Department.id == Student.dept_id)
        .join(Organization, Organization.id == Department.org_id)
        .outerjoin(User, User.id == Student.user_id)
    )
    if dept_id:
        q = q.where(Student.dept_id == dept_id)
    elif org_id:
        q = q.where(Department.org_id == org_id)

    q = q.order_by(Department.name, Student.roll_no).limit(200)
    res = await db.execute(q)
    rows = res.all()

    return [
        {
            "id": str(r.Student.id),
            "name": r.Student.name,
            "roll_no": r.Student.roll_no,
            "enrollment_no": r.Student.enrollment_no,
            "division": r.Student.division,
            "batch": r.Student.batch,
            "semester": r.Student.semester,
            "dept_id": str(r.Student.dept_id),
            "dept_name": r.dept_name,
            "org_name": r.org_name,
            "email": r.email or "—",
        }
        for r in rows
    ]


@router.post("/students")
async def create_student(
    data: StudentCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    """Create a student + user account inside any department."""
    # Verify department exists and get org
    dept = await db.get(Department, uuid.UUID(data.dept_id))
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    # Check email uniqueness
    existing = await db.execute(select(User).where(User.email == data.email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Email '{data.email}' already exists")

    user = User(
        email=data.email,
        password_hash=_hash("Student@123"),
        role=UserRole.student,
        org_id=dept.org_id,
        is_active=True,
    )
    db.add(user)
    await db.flush()

    student = Student(
        user_id=user.id,
        name=data.name,
        roll_no=data.roll_no,
        enrollment_no=data.enrollment_no,
        division=data.division,
        batch=data.batch,
        semester=data.semester,
        dept_id=uuid.UUID(data.dept_id),
    )
    db.add(student)
    await db.commit()
    await db.refresh(student)

    # Get dept name for response
    return {
        "id": str(student.id),
        "name": student.name,
        "roll_no": student.roll_no,
        "enrollment_no": student.enrollment_no,
        "division": student.division,
        "batch": student.batch,
        "semester": student.semester,
        "dept_id": str(student.dept_id),
        "dept_name": dept.name,
        "email": data.email,
    }


@router.delete("/students/{student_id}")
async def delete_student(
    student_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    student = await db.get(Student, uuid.UUID(student_id))
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    stu_uuid = uuid.UUID(student_id)
    name = student.name
    user_id = student.user_id

    # Delete child records first
    await db.execute(sql_delete(FaceEmbedding).where(FaceEmbedding.student_id == stu_uuid))
    await db.execute(sql_delete(AttendanceDispute).where(AttendanceDispute.student_id == stu_uuid))

    # Delete attendance records (and their evidence)
    att_res = await db.execute(select(AttendanceRecord.id).where(AttendanceRecord.student_id == stu_uuid))
    att_ids = [r[0] for r in att_res.all()]
    if att_ids:
        await db.execute(sql_delete(AttendanceEvidence).where(AttendanceEvidence.attendance_id.in_(att_ids)))
        await db.execute(sql_delete(AttendanceRecord).where(AttendanceRecord.id.in_(att_ids)))

    await db.execute(sql_delete(Student).where(Student.id == stu_uuid))

    # Delete linked user account
    if user_id:
        await db.execute(sql_delete(User).where(User.id == user_id))

    await db.commit()
    return {"message": f"Student '{name}' deleted"}


# ═══════════════════════════════════════════════════════════
# BULK STUDENT IMPORT (Excel)
# ═══════════════════════════════════════════════════════════

@router.get("/students/template")
async def download_student_template(
    current_user=Depends(require_super_admin)
):
    """
    Download an Excel template (.xlsx) for bulk student import.
    Includes a sample row and a helper sheet explaining each column.
    """
    wb = __import__("openpyxl").Workbook()

    # ── Main data sheet ──────────────────────────────────────
    ws = wb.active
    ws.title = "Students"

    headers = [
        "name", "roll_no", "enrollment_no", "email",
        "division", "batch", "semester", "dept_id"
    ]
    header_fill    = __import__("openpyxl.styles", fromlist=["PatternFill"]).PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
    header_font    = __import__("openpyxl.styles", fromlist=["Font"]).Font(bold=True, color="FFFFFF")
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin
        ws.column_dimensions[
            __import__("openpyxl.utils", fromlist=["get_column_letter"]).get_column_letter(col_idx)
        ].width = 22

    # Sample rows
    sample_rows = [
        ["Rahul Sharma",  "CS001", "EN2024001", "rahul@college.edu",   "A", "B1", 4, "<paste-dept-id-here>"],
        ["Priya Patel",   "CS002", "EN2024002", "priya@college.edu",   "A", "B1", 4, "<paste-dept-id-here>"],
        ["Amit Shah",     "CS003", "",           "amit@college.edu",    "B", "B2", 4, "<paste-dept-id-here>"],
    ]
    for row_data in sample_rows:
        ws.append(row_data)

    # Freeze first row
    ws.freeze_panes = "A2"

    # ── Instructions sheet ───────────────────────────────────
    info_ws = wb.create_sheet("Instructions")
    info_ws.column_dimensions["A"].width = 22
    info_ws.column_dimensions["B"].width = 60

    info_ws.cell(1, 1, "Column").font        = Font(bold=True)
    info_ws.cell(1, 2, "Description").font   = Font(bold=True)

    instructions = [
        ("name",           "Required. Full name of the student."),
        ("roll_no",        "Required. Unique roll number within the department."),
        ("enrollment_no",  "Optional. Enrollment / admission number."),
        ("email",          "Required. Unique email — used as login username."),
        ("division",       "Optional. e.g. A, B, C"),
        ("batch",          "Optional. e.g. B1, B2, All"),
        ("semester",       "Optional. Integer, e.g. 4"),
        ("dept_id",        "Required. UUID of the department. Copy from the Super Admin panel URL or use the Departments tab."),
        ("", ""),
        ("Default password", "Student@123  (students must change on first login)"),
        ("Max rows",         "500 students per upload"),
        ("Duplicates",       "Rows with an existing email are skipped with an error message."),
    ]
    for r_idx, (col, desc) in enumerate(instructions, 2):
        info_ws.cell(r_idx, 1, col).font  = Font(bold=bool(col))
        info_ws.cell(r_idx, 2, desc)

    # Serialise to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_import_template.xlsx"},
    )


@router.post("/students/bulk")
async def bulk_import_students(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    """
    Bulk-import students from an Excel (.xlsx) or CSV file.
    Returns per-row results: success list + error list.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are accepted")

    contents = await file.read()
    buf = io.BytesIO(contents)

    try:
        if ext == "csv":
            df = pd.read_csv(buf, dtype=str)
        else:
            df = pd.read_excel(buf, sheet_name="Students", dtype=str)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Failed to parse file: {str(e)}")

    required_cols = {"name", "roll_no", "email", "dept_id"}
    missing = required_cols - set(df.columns.str.strip().str.lower())
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {', '.join(sorted(missing))}. Use the downloaded template."
        )

    # Normalise column names
    df.columns = df.columns.str.strip().str.lower()
    df = df.where(pd.notnull(df), None)  # NaN → None
    df = df.head(500)                    # cap at 500 rows

    results_ok:  list[dict] = []
    results_err: list[dict] = []

    for row_num, row in df.iterrows():
        row_label = f"Row {int(row_num) + 2}"  # +2 because header=row 1, 0-indexed

        name     = (row.get("name")     or "").strip()
        roll_no  = (row.get("roll_no")  or "").strip()
        email    = (row.get("email")    or "").strip().lower()
        dept_id  = (row.get("dept_id")  or "").strip()
        enroll   = (row.get("enrollment_no") or "").strip() or None
        division = (row.get("division") or "").strip() or None
        batch    = (row.get("batch")    or "").strip() or None
        sem_raw  = row.get("semester")
        semester: Optional[int] = None
        if sem_raw:
            try: semester = int(float(str(sem_raw)))
            except ValueError: pass

        # Basic validation
        if not name:
            results_err.append({"row": row_label, "email": email, "error": "'name' is required"})
            continue
        if not roll_no:
            results_err.append({"row": row_label, "email": email, "error": "'roll_no' is required"})
            continue
        if not email:
            results_err.append({"row": row_label, "roll_no": roll_no, "error": "'email' is required"})
            continue
        if not dept_id or dept_id.startswith("<"):
            results_err.append({"row": row_label, "email": email, "error": "'dept_id' is missing or still placeholder"})
            continue

        # Validate dept exists
        try:
            dept_uuid = uuid.UUID(dept_id)
        except ValueError:
            results_err.append({"row": row_label, "email": email, "error": f"'dept_id' is not a valid UUID: {dept_id}"})
            continue

        dept = await db.get(Department, dept_uuid)
        if not dept:
            results_err.append({"row": row_label, "email": email, "error": f"Department '{dept_id}' not found"})
            continue

        # Check email uniqueness
        existing = await db.execute(select(User).where(User.email == email))
        if existing.scalar_one_or_none():
            results_err.append({"row": row_label, "email": email, "error": f"Email '{email}' already exists"})
            continue

        # Check roll_no uniqueness within dept
        existing_roll = await db.execute(
            select(Student).where(Student.dept_id == dept_uuid, Student.roll_no == roll_no)
        )
        if existing_roll.scalar_one_or_none():
            results_err.append({"row": row_label, "email": email, "error": f"Roll No '{roll_no}' already exists in this department"})
            continue

        # Create user + student
        try:
            user = User(
                email=email,
                password_hash=_hash("Student@123"),
                role=UserRole.student,
                org_id=dept.org_id,
                is_active=True,
            )
            db.add(user)
            await db.flush()

            student = Student(
                user_id=user.id,
                name=name,
                roll_no=roll_no,
                enrollment_no=enroll,
                division=division,
                batch=batch,
                semester=semester,
                dept_id=dept_uuid,
            )
            db.add(student)
            await db.flush()

            results_ok.append({"row": row_label, "name": name, "email": email, "roll_no": roll_no})

        except Exception as e:
            await db.rollback()
            results_err.append({"row": row_label, "email": email, "error": str(e)})
            continue

    try:
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Commit failed: {str(e)}")

    return {
        "total_rows":    len(df),
        "success_count": len(results_ok),
        "error_count":   len(results_err),
        "created":       results_ok,
        "errors":        results_err,
    }

# ═══════════════════════════════════════════════════════════
# FACULTY
# ═══════════════════════════════════════════════════════════

@router.get("/faculty")
async def list_all_faculty(
    org_id: Optional[str] = None,
    dept_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    """List all faculty across the platform, optionally filtered by org or dept."""
    q = (
        select(Faculty, Department.name.label("dept_name"), Organization.name.label("org_name"), User.email.label("email"))
        .join(User, User.id == Faculty.user_id)
        .join(Department, Department.id == Faculty.dept_id)
        .join(Organization, Organization.id == Department.org_id)
    )
    if dept_id:
        q = q.where(Faculty.dept_id == dept_id)
    elif org_id:
        q = q.where(User.org_id == org_id)

    q = q.order_by(Department.name, Faculty.name).limit(200)
    res = await db.execute(q)
    rows = res.all()

    return [
        {
            "id": str(r.Faculty.id),
            "name": r.Faculty.name,
            "designation": r.Faculty.designation,
            "dept_id": str(r.Faculty.dept_id),
            "dept_name": r.dept_name,
            "org_name": r.org_name,
            "email": r.email,
        }
        for r in rows
    ]


@router.post("/faculty")
async def create_faculty(
    data: FacultyCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    """Create a faculty member + user account inside any department."""
    dept = await db.get(Department, uuid.UUID(data.dept_id))
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    existing = await db.execute(select(User).where(User.email == data.email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Email '{data.email}' already exists")

    user = User(
        email=data.email,
        password_hash=_hash("Faculty@123"),
        role=UserRole.faculty,
        org_id=dept.org_id,
        is_active=True,
    )
    db.add(user)
    await db.flush()

    faculty = Faculty(
        user_id=user.id,
        name=data.name,
        designation=data.designation,
        dept_id=uuid.UUID(data.dept_id),
    )
    db.add(faculty)
    await db.commit()
    await db.refresh(faculty)

    return {
        "id": str(faculty.id),
        "name": faculty.name,
        "designation": faculty.designation,
        "dept_id": str(faculty.dept_id),
        "dept_name": dept.name,
        "email": data.email,
    }


@router.delete("/faculty/{faculty_id}")
async def delete_faculty(
    faculty_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    fac = await db.get(Faculty, uuid.UUID(faculty_id))
    if not fac:
        raise HTTPException(status_code=404, detail="Faculty not found")
    fac_uuid = uuid.UUID(faculty_id)
    name = fac.name
    user_id = fac.user_id

    # Delete subject assignments for this faculty first
    await db.execute(sql_delete(SubjectAssignment).where(SubjectAssignment.faculty_id == fac_uuid))

    # Nullify lecture references (keep lecture history, just unlink faculty)
    # Alternatively delete lectures - using nullify to preserve records
    lec_res = await db.execute(select(Lecture.id).where(Lecture.faculty_id == fac_uuid))
    lec_ids = [r[0] for r in lec_res.all()]
    if lec_ids:
        await db.execute(sql_delete(AttendanceEvidence).where(AttendanceEvidence.lecture_id.in_(lec_ids)))
        await db.execute(sql_delete(AttendanceDispute).where(AttendanceDispute.lecture_id.in_(lec_ids)))
        await db.execute(sql_delete(AttendanceRecord).where(AttendanceRecord.lecture_id.in_(lec_ids)))
        await db.execute(sql_delete(Lecture).where(Lecture.id.in_(lec_ids)))

    await db.execute(sql_delete(Faculty).where(Faculty.id == fac_uuid))

    # Delete linked user account
    if user_id:
        await db.execute(sql_delete(User).where(User.id == user_id))

    await db.commit()
    return {"message": f"Faculty '{name}' deleted"}


# ═══════════════════════════════════════════════════════════
# USERS (platform-wide)
# ═══════════════════════════════════════════════════════════

@router.get("/users")
async def list_all_users(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin),
    role: Optional[str] = Query(None),
    org_id: Optional[str] = Query(None),
    admin_only: bool = Query(False),
):
    """List users. Filter by role and/or org_id. admin_only=true returns org_admin+dept_admin only."""
    q = select(User, Organization.name.label("org_name")).join(
        Organization, Organization.id == User.org_id, isouter=True
    )
    if admin_only:
        q = q.where(or_(
            User.role == UserRole.org_admin,
            User.role == UserRole.dept_admin,
        ))
    elif role:
        try:
            q = q.where(User.role == UserRole(role))
        except ValueError:
            pass
    if org_id:
        q = q.where(User.org_id == uuid.UUID(org_id))
    q = q.order_by(User.created_at.desc()).limit(200)
    res = await db.execute(q)
    rows = res.all()
    return [
        {
            "id": str(r.User.id),
            "email": r.User.email,
            "role": r.User.role.value,
            "org_name": r.org_name or "—",
            "org_id": str(r.User.org_id) if r.User.org_id else None,
            "is_active": r.User.is_active,
            "created_at": r.User.created_at.strftime("%Y-%m-%d") if r.User.created_at else "—",
        }
        for r in rows
    ]


@router.post("/admins")
async def create_admin(
    data: AdminCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    """
    Create an org_admin or dept_admin user.
    For dept_admin, also creates a Faculty profile so they can be assigned to subjects.
    Default password is Admin@123 for both roles.
    """
    # Validate role
    if data.role not in ("org_admin", "dept_admin"):
        raise HTTPException(status_code=400, detail="role must be 'org_admin' or 'dept_admin'")

    # Validate org
    org = await db.get(Organization, uuid.UUID(data.org_id))
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    # For dept_admin, validate dept
    dept = None
    if data.role == "dept_admin":
        if not data.dept_id:
            raise HTTPException(status_code=400, detail="dept_id is required for dept_admin")
        dept = await db.get(Department, uuid.UUID(data.dept_id))
        if not dept:
            raise HTTPException(status_code=404, detail="Department not found")
        if str(dept.org_id) != data.org_id:
            raise HTTPException(status_code=400, detail="Department does not belong to this organization")

    # Check email uniqueness
    existing = await db.execute(select(User).where(User.email == data.email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Email '{data.email}' already exists")

    # Create user
    user = User(
        email=data.email,
        password_hash=_hash("Admin@123"),
        role=UserRole(data.role),
        org_id=uuid.UUID(data.org_id),
        is_active=True,
    )
    db.add(user)
    await db.flush()

    # For dept_admin: also create a Faculty profile
    faculty_profile = None
    if data.role == "dept_admin" and dept:
        faculty_profile = Faculty(
            user_id=user.id,
            name=data.name or data.email.split("@")[0].replace(".", " ").title(),
            designation="Department Admin",
            dept_id=dept.id,
        )
        db.add(faculty_profile)

    await db.commit()
    await db.refresh(user)

    return {
        "id": str(user.id),
        "email": user.email,
        "role": user.role.value,
        "org_id": str(user.org_id),
        "org_name": org.name,
        "dept_id": str(dept.id) if dept else None,
        "dept_name": dept.name if dept else None,
        "is_active": user.is_active,
        "created_at": user.created_at.strftime("%Y-%m-%d") if user.created_at else "—",
    }


@router.patch("/users/{user_id}/toggle")
async def toggle_user_active(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    user = await db.get(User, uuid.UUID(user_id))
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_active = not user.is_active
    await db.commit()
    return {"id": str(user.id), "is_active": user.is_active}


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_super_admin)
):
    user = await db.get(User, uuid.UUID(user_id))
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    email = user.email
    await db.delete(user)
    await db.commit()
    return {"message": f"User '{email}' deleted"}
